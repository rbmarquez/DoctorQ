# src/utils/excel_processor_hybrid.py

import asyncio
import hashlib
import json
import logging
import os
import time
import warnings
from pathlib import Path
from typing import Any, Dict, Optional

import pandas as pd
import psutil

# Suprimir warnings do pandas e openpyxl
warnings.filterwarnings(
    "ignore", message=".*Data Validation extension is not supported.*"
)
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

logger = logging.getLogger(__name__)


class HybridExcelProcessor:
    """
    Processador hÃ­brido inteligente para arquivos Excel de todos os tamanhos.

    EstratÃ©gias:
    1. Excel pequenos (â‰¤ 3 abas, â‰¤ 1K linhas): openpyxl rÃ¡pido
    2. Excel mÃ©dios (â‰¤ 10 abas, â‰¤ 10K linhas): pandas com chunks
    3. Excel grandes (â‰¤ 50 abas): processamento aba por aba
    4. Excel gigantes (> 50 abas): processamento assÃ­ncrono obrigatÃ³rio
    """

    # Limites para classificaÃ§Ã£o automÃ¡tica
    SMALL_EXCEL_MAX_SHEETS = 3
    SMALL_EXCEL_MAX_ROWS = 1000

    MEDIUM_EXCEL_MAX_SHEETS = 10
    MEDIUM_EXCEL_MAX_ROWS = 10000

    LARGE_EXCEL_MAX_SHEETS = 50

    # ConfiguraÃ§Ãµes de performance otimizadas para arquivos grandes
    CHUNK_SIZE_ROWS = 3000  # Processar em chunks menores para arquivos grandes
    MAX_SHEETS_PER_CHUNK = 3  # MÃ¡ximo 3 abas por chunk para arquivos grandes
    CACHE_EXPIRY_HOURS = 24
    MAX_PROCESSING_TIME = 900  # 15 minutos para arquivos muito grandes
    MAX_MEMORY_MB = 3072  # 3GB limite de memÃ³ria

    def __init__(self):
        self.logger = logger
        self._cache_dir = self._setup_cache_dir()
        self._memory_limit_mb = self.MAX_MEMORY_MB

    def _setup_cache_dir(self) -> Path:
        """Configura diretÃ³rio de cache"""
        # Detectar ambiente K8s
        is_k8s = os.getenv("KUBERNETES_SERVICE_HOST") is not None or os.path.exists(
            "/.dockerenv"
        )

        if is_k8s:
            # Em K8s, usar /tmp que Ã© garantido ter permissÃµes de escrita
            cache_dir = Path("/tmp") / "inovaia_excel_cache"
            self.logger.debug(f"Usando cache Excel K8s: {cache_dir}")
        else:
            # Em ambiente local, usar diretÃ³rio do projeto
            project_root = Path(__file__).parent.parent.parent
            cache_dir = project_root / "temp" / "excel_cache"
            self.logger.debug(f"Usando cache Excel local: {cache_dir}")

        cache_dir.mkdir(parents=True, exist_ok=True)
        return cache_dir

    def _get_file_hash(self, file_path: str) -> str:
        """Gera hash Ãºnico do arquivo para cache"""
        # Verificar se o arquivo existe e nÃ£o estÃ¡ vazio
        if not os.path.exists(file_path):
            self.logger.error(f"Arquivo Excel nÃ£o existe para hash: {file_path}")
            raise FileNotFoundError(f"Arquivo nÃ£o encontrado: {file_path}")

        file_size = os.path.getsize(file_path)
        if file_size == 0:
            self.logger.error(f"Arquivo Excel estÃ¡ vazio para hash: {file_path}")
            raise ValueError(f"Arquivo estÃ¡ vazio: {file_path}")

        with open(file_path, "rb") as f:
            file_hash = hashlib.md5(f.read()).hexdigest()
        return file_hash

    def _get_memory_usage_mb(self) -> float:
        """Retorna uso de memÃ³ria atual em MB"""
        try:
            process = psutil.Process(os.getpid())
            return process.memory_info().rss / 1024 / 1024
        except Exception:
            return 0

    def _check_memory_limit(self) -> bool:
        """Verifica se hÃ¡ memÃ³ria suficiente disponÃ­vel"""
        current_memory = self._get_memory_usage_mb()
        return current_memory < (self._memory_limit_mb * 0.8)  # 80% do limite

    async def _estimate_excel_complexity(self, file_path: str) -> Dict[str, Any]:
        """Estima complexidade e caracterÃ­sticas do arquivo Excel"""
        try:
            file_size_mb = os.path.getsize(file_path) / 1024 / 1024

            # Usar pandas para anÃ¡lise rÃ¡pida das abas
            excel_file = pd.ExcelFile(file_path)
            sheet_names = excel_file.sheet_names
            sheet_count = len(sheet_names)

            # Analisar primeira aba para estimar complexidade
            total_rows = 0
            total_cols = 0
            has_formulas = False
            has_merged_cells = False

            # Para arquivos grandes, analisar menos abas mais rapidamente
            sheets_to_analyze = min(2 if file_size_mb > 50 else 3, sheet_count)

            for sheet_name in sheet_names[:sheets_to_analyze]:
                try:
                    # Para arquivos grandes, ler menos linhas para anÃ¡lise
                    nrows_sample = 50 if file_size_mb > 50 else 100
                    df = pd.read_excel(
                        file_path, sheet_name=sheet_name, nrows=nrows_sample
                    )

                    # Para arquivos grandes, usar mÃ©todo mais rÃ¡pido de contagem
                    if file_size_mb > 50:
                        rows_in_sheet, cols_in_sheet = (
                            self._estimate_large_file_dimensions(
                                file_path, sheet_name, file_size_mb, sheet_count
                            )
                        )
                    else:
                        # Contar todas as linhas da aba (mÃ©todo original para arquivos menores)
                        full_df = pd.read_excel(file_path, sheet_name=sheet_name)
                        rows_in_sheet = len(full_df)
                        cols_in_sheet = len(full_df.columns)

                    total_rows += rows_in_sheet
                    total_cols = max(total_cols, cols_in_sheet)

                    # Verificar se hÃ¡ cÃ©lulas com fÃ³rmulas (heurÃ­stica)
                    if any(
                        str(cell).startswith("=")
                        for col in df.columns
                        for cell in df[col].astype(str)[:10]
                    ):
                        has_formulas = True

                except Exception as e:
                    self.logger.warning(f"Erro ao analisar aba '{sheet_name}': {e}")
                    # Estimativa conservadora
                    total_rows += 1000
                    total_cols = max(total_cols, 10)

            # ExtrapolaÃ§Ã£o para abas nÃ£o analisadas
            if sheet_count > sheets_to_analyze:
                avg_rows_per_sheet = total_rows / sheets_to_analyze
                remaining_sheets = sheet_count - sheets_to_analyze
                total_rows += int(avg_rows_per_sheet * remaining_sheets)

            excel_file.close()

            complexity = {
                "sheet_count": sheet_count,
                "total_rows": total_rows,
                "max_columns": total_cols,
                "file_size_mb": file_size_mb,
                "has_formulas": has_formulas,
                "has_merged_cells": has_merged_cells,
                "estimated_processing_time": self._estimate_processing_time(
                    sheet_count, total_rows, file_size_mb, has_formulas
                ),
                "sheet_names": (
                    sheet_names[:10] if len(sheet_names) > 10 else sheet_names
                ),  # Top 10 para debug
            }

            self.logger.info(f"Excel Complexity Analysis: {complexity}")
            return complexity

        except Exception as e:
            self.logger.error(f"Erro ao analisar Excel: {e}")
            # Fallback: estimar apenas por tamanho do arquivo
            file_size_mb = os.path.getsize(file_path) / 1024 / 1024
            estimated_sheets = max(1, min(int(file_size_mb * 2), 20))  # ~0.5MB por aba
            estimated_rows = max(100, int(file_size_mb * 500))  # ~2KB por linha

            return {
                "sheet_count": estimated_sheets,
                "total_rows": estimated_rows,
                "max_columns": 20,  # Estimativa conservadora
                "file_size_mb": file_size_mb,
                "has_formulas": True,  # Assumir complexidade mÃ©dia
                "has_merged_cells": False,
                "estimated_processing_time": self._estimate_processing_time(
                    estimated_sheets, estimated_rows, file_size_mb, True
                ),
                "sheet_names": ["Sheet1", "Sheet2", "Sheet3"],  # PadrÃ£o
            }

    def _estimate_large_file_dimensions(
        self, file_path: str, sheet_name: str, file_size_mb: float, sheet_count: int
    ) -> tuple[int, int]:
        """Estima dimensÃµes de arquivos grandes"""
        try:
            # Tentar ler sÃ³ o header para contar colunas
            header_df = pd.read_excel(file_path, sheet_name=sheet_name, nrows=0)
            cols_in_sheet = len(header_df.columns)

            # Estimar linhas baseado no tamanho do arquivo e nÃºmero de abas
            estimated_rows_per_sheet = max(
                1000, int((file_size_mb * 1000) / sheet_count)
            )
            rows_in_sheet = estimated_rows_per_sheet
        except Exception:
            # Fallback: contar todas as linhas (mais lento)
            full_df = pd.read_excel(file_path, sheet_name=sheet_name)
            rows_in_sheet = len(full_df)
            cols_in_sheet = len(full_df.columns)

        return rows_in_sheet, cols_in_sheet

    def _estimate_processing_time(
        self, sheets: int, rows: int, size_mb: float, has_formulas: bool
    ) -> int:
        """Estima tempo de processamento em segundos"""
        base_time_per_row = 0.001  # 1ms por linha para dados simples
        if has_formulas:
            base_time_per_row *= 2  # 2x mais lento com fÃ³rmulas

        base_time_per_sheet = 0.5  # 0.5s por aba

        estimated_seconds = (rows * base_time_per_row) + (sheets * base_time_per_sheet)

        # Adicionar overhead por tamanho do arquivo
        size_overhead = size_mb * 0.2  # 0.2s por MB

        return int(estimated_seconds + size_overhead)

    def _select_strategy(self, complexity: Dict[str, Any]) -> str:
        """Seleciona estratÃ©gia de processamento baseada na complexidade"""
        sheets = complexity["sheet_count"]
        total_rows = complexity["total_rows"]
        size_mb = complexity["file_size_mb"]
        has_formulas = complexity["has_formulas"]

        # EstratÃ©gia 1: Excel pequeno e simples -> openpyxl rÃ¡pido
        if (
            sheets <= self.SMALL_EXCEL_MAX_SHEETS
            and total_rows <= self.SMALL_EXCEL_MAX_ROWS
            and size_mb <= 5
            and not has_formulas
        ):
            return "openpyxl_fast"

        # EstratÃ©gia 2: Excel mÃ©dio -> pandas com chunks
        if (
            sheets <= self.MEDIUM_EXCEL_MAX_SHEETS
            and total_rows <= self.MEDIUM_EXCEL_MAX_ROWS
            and size_mb <= 80  # Aumentado para 80MB
        ):
            return "pandas_chunked"

        # EstratÃ©gia 3: Excel grande -> processamento aba por aba
        if (
            sheets <= self.LARGE_EXCEL_MAX_SHEETS and size_mb <= 300
        ):  # Aumentado para 300MB
            return "sheet_by_sheet"

        # EstratÃ©gia 4: Excel gigante -> processamento assÃ­ncrono obrigatÃ³rio
        return "async_required"

    async def _extract_with_openpyxl_fast(self, file_path: str) -> str:
        """ExtraÃ§Ã£o ultra rÃ¡pida com openpyxl para Excel pequenos"""
        try:
            from openpyxl import load_workbook

            wb = load_workbook(file_path, data_only=True, read_only=True)
            text_parts = []

            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_text = f"=== ABA: {sheet_name} ===\n"

                # Ler todas as cÃ©lulas com dados
                rows_with_data = []
                for row in ws.iter_rows(values_only=True):
                    # Filtrar linhas vazias
                    row_data = [str(cell) if cell is not None else "" for cell in row]
                    if any(cell.strip() for cell in row_data):
                        rows_with_data.append("\t".join(row_data))

                if rows_with_data:
                    sheet_text += "\n".join(rows_with_data)
                    text_parts.append(sheet_text)

            wb.close()
            return "\n\n".join(text_parts)

        except Exception as e:
            self.logger.error(f"Erro openpyxl Fast: {e}")
            raise

    async def _extract_with_pandas_chunked(self, file_path: str) -> str:
        """ExtraÃ§Ã£o com pandas usando chunks para Excel mÃ©dios"""
        try:
            excel_file = pd.ExcelFile(file_path)
            text_parts = []

            for sheet_name in excel_file.sheet_names:
                sheet_text = f"=== ABA: {sheet_name} ===\n"

                try:
                    # Ler aba em chunks se for muito grande
                    df = pd.read_excel(file_path, sheet_name=sheet_name)

                    if len(df) > self.CHUNK_SIZE_ROWS:
                        # Processar em chunks
                        chunk_texts = []
                        for i in range(0, len(df), self.CHUNK_SIZE_ROWS):
                            chunk = df.iloc[i : i + self.CHUNK_SIZE_ROWS]
                            chunk_text = chunk.to_csv(sep="\t", index=False, na_rep="")
                            chunk_texts.append(chunk_text)

                            # Verificar memÃ³ria
                            if not self._check_memory_limit():
                                self.logger.warning(
                                    f"Limite de memÃ³ria atingido na aba {sheet_name}"
                                )
                                break

                        sheet_text += "\n".join(chunk_texts)
                    else:
                        # Processar aba inteira
                        sheet_text += df.to_csv(sep="\t", index=False, na_rep="")

                    text_parts.append(sheet_text)

                except Exception as e:
                    self.logger.warning(f"Erro ao processar aba '{sheet_name}': {e}")
                    text_parts.append(
                        f"=== ABA: {sheet_name} ===\n[ERRO AO PROCESSAR ABA: {e}]"
                    )

            excel_file.close()
            return "\n\n".join(text_parts)

        except Exception as e:
            self.logger.error(f"Erro pandas chunked: {e}")
            raise

    async def _extract_sheet_by_sheet(self, file_path: str) -> str:
        """Processamento aba por aba para Excel grandes"""
        try:
            excel_file = pd.ExcelFile(file_path)
            text_parts = []
            processed_sheets = 0

            # Processar em grupos de abas para controlar memÃ³ria
            sheet_names = excel_file.sheet_names

            for i in range(0, len(sheet_names), self.MAX_SHEETS_PER_CHUNK):
                # Verificar memÃ³ria antes de cada chunk
                if not self._check_memory_limit():
                    self.logger.warning(
                        "Limite de memÃ³ria atingido, interrompendo processamento"
                    )
                    break

                chunk_sheets = sheet_names[i : i + self.MAX_SHEETS_PER_CHUNK]
                self.logger.info(
                    f"Processando abas {i+1}-{i+len(chunk_sheets)} de {len(sheet_names)}"
                )

                for sheet_name in chunk_sheets:
                    try:
                        df = pd.read_excel(file_path, sheet_name=sheet_name)
                        sheet_text = f"=== ABA: {sheet_name} ===\n"

                        # Limitar dados se a aba for muito grande
                        if len(df) > 50000:  # 50K linhas
                            sheet_text += (
                                f"[ABA GRANDE - PRIMEIRAS 50.000 LINHAS DE {len(df)}]\n"
                            )
                            df = df.head(50000)

                        sheet_text += df.to_csv(sep="\t", index=False, na_rep="")
                        text_parts.append(sheet_text)
                        processed_sheets += 1

                        # Pequena pausa para evitar sobrecarga
                        await asyncio.sleep(0.1)

                    except Exception as e:
                        self.logger.warning(
                            f"Erro ao processar aba '{sheet_name}': {e}"
                        )
                        text_parts.append(
                            f"=== ABA: {sheet_name} ===\n[ERRO AO PROCESSAR: {e}]"
                        )

                # Limpeza de memÃ³ria entre chunks
                import gc

                gc.collect()

            excel_file.close()

            if processed_sheets < len(sheet_names):
                text_parts.append(
                    f"\n[PROCESSAMENTO INTERROMPIDO - {processed_sheets} de {len(sheet_names)} abas processadas]"
                )

            return "\n\n".join(text_parts)

        except Exception as e:
            self.logger.error(f"Erro sheet by sheet: {e}")
            raise

    def _save_to_cache(self, file_hash: str, content: str, strategy: str):
        """Salva resultado no cache"""
        try:
            cache_data = {
                "content": content,
                "strategy": strategy,
                "timestamp": time.time(),
                "size": len(content),
            }

            cache_file = self._cache_dir / f"{file_hash}.json"
            with open(cache_file, "w", encoding="utf-8") as f:
                json.dump(cache_data, f, ensure_ascii=False, indent=2)

            self.logger.debug(f"Resultado Excel salvo no cache: {cache_file}")

        except Exception as e:
            self.logger.warning(f"Erro ao salvar cache Excel: {e}")

    def _load_from_cache(self, file_hash: str) -> Optional[str]:
        """Carrega resultado do cache se vÃ¡lido"""
        try:
            cache_file = self._cache_dir / f"{file_hash}.json"

            if not cache_file.exists():
                return None

            with open(cache_file, "r", encoding="utf-8") as f:
                cache_data = json.load(f)

            # Verificar se cache ainda Ã© vÃ¡lido
            cache_age_hours = (time.time() - cache_data["timestamp"]) / 3600
            if cache_age_hours > self.CACHE_EXPIRY_HOURS:
                # Cache expirado, remover
                cache_file.unlink()
                return None

            self.logger.info(
                f"Resultado Excel carregado do cache (idade: {cache_age_hours:.1f}h)"
            )
            return cache_data["content"]

        except Exception as e:
            self.logger.warning(f"Erro ao carregar cache Excel: {e}")
            return None

    async def process_excel(
        self, file_path: str, use_cache: bool = True
    ) -> Dict[str, Any]:
        """
        Processa Excel usando estratÃ©gia hÃ­brida inteligente

        Returns:
            Dict com conteÃºdo extraÃ­do e metadados do processamento
        """
        start_time = time.time()

        try:
            # 1. Verificar cache se habilitado
            file_hash = self._get_file_hash(file_path) if use_cache else None

            if use_cache and file_hash:
                cached_content = self._load_from_cache(file_hash)
                if cached_content:
                    return {
                        "content": cached_content,
                        "processing_info": {
                            "strategy": "cache_hit",
                            "processing_time": time.time() - start_time,
                            "from_cache": True,
                            "file_hash": file_hash,
                        },
                    }

            # 2. Analisar complexidade do Excel
            complexity = await self._estimate_excel_complexity(file_path)

            # 3. Selecionar estratÃ©gia
            strategy = self._select_strategy(complexity)

            self.logger.info(
                f"EstratÃ©gia Excel selecionada: {strategy} para arquivo com {complexity['sheet_count']} abas e {complexity['total_rows']} linhas"
            )

            # 4. Verificar se Ã© processamento assÃ­ncrono obrigatÃ³rio
            if strategy == "async_required":
                raise ValueError(
                    f"Excel muito grande ({complexity['sheet_count']} abas, {complexity['total_rows']} linhas, {complexity['file_size_mb']:.1f}MB). "
                    f"Processamento assÃ­ncrono obrigatÃ³rio. Tempo estimado: {complexity['estimated_processing_time']}s"
                )

            # 5. Processar com timeout
            content = await asyncio.wait_for(
                self._execute_strategy(strategy, file_path),
                timeout=self.MAX_PROCESSING_TIME,
            )

            processing_time = time.time() - start_time

            # 6. Salvar no cache
            if use_cache and file_hash:
                self._save_to_cache(file_hash, content, strategy)

            # 7. Retornar resultado
            return {
                "content": content,
                "processing_info": {
                    "strategy": strategy,
                    "processing_time": processing_time,
                    "complexity": complexity,
                    "from_cache": False,
                    "file_hash": file_hash,
                    "content_size": len(content),
                },
            }

        except asyncio.TimeoutError:
            raise TimeoutError(
                f"Processamento Excel excedeu {self.MAX_PROCESSING_TIME}s. "
                "Use processamento assÃ­ncrono para arquivos muito grandes."
            )
        except Exception as e:
            self.logger.error(f"Erro no processamento hÃ­brido Excel: {e}")
            raise

    async def _execute_strategy(self, strategy: str, file_path: str) -> str:
        """Executa a estratÃ©gia selecionada"""
        if strategy == "openpyxl_fast":
            return await self._extract_with_openpyxl_fast(file_path)
        if strategy == "pandas_chunked":
            return await self._extract_with_pandas_chunked(file_path)
        if strategy == "sheet_by_sheet":
            return await self._extract_sheet_by_sheet(file_path)

        raise ValueError(f"EstratÃ©gia Excel nÃ£o implementada: {strategy}")

    def get_strategy_info(self) -> Dict[str, Any]:
        """Retorna informaÃ§Ãµes sobre as estratÃ©gias disponÃ­veis"""
        return {
            "strategies": {
                "openpyxl_fast": "Ultra rÃ¡pido para Excel pequenos (â‰¤3 abas, â‰¤1K linhas, â‰¤5MB, sem fÃ³rmulas)",
                "pandas_chunked": "Chunks otimizados para Excel mÃ©dios (â‰¤10 abas, â‰¤10K linhas, â‰¤50MB)",
                "sheet_by_sheet": "Processamento aba por aba para Excel grandes (â‰¤50 abas, â‰¤200MB)",
                "async_required": "Processamento assÃ­ncrono obrigatÃ³rio (>50 abas ou >200MB)",
            },
            "limits": {
                "small_excel_max_sheets": self.SMALL_EXCEL_MAX_SHEETS,
                "small_excel_max_rows": self.SMALL_EXCEL_MAX_ROWS,
                "medium_excel_max_sheets": self.MEDIUM_EXCEL_MAX_SHEETS,
                "medium_excel_max_rows": self.MEDIUM_EXCEL_MAX_ROWS,
                "large_excel_max_sheets": self.LARGE_EXCEL_MAX_SHEETS,
                "chunk_size_rows": self.CHUNK_SIZE_ROWS,
                "max_sheets_per_chunk": self.MAX_SHEETS_PER_CHUNK,
                "max_processing_time": self.MAX_PROCESSING_TIME,
            },
            "features": {
                "intelligent_strategy_selection": True,
                "automatic_caching": True,
                "memory_monitoring": True,
                "chunked_processing": True,
                "sheet_by_sheet_processing": True,
                "formula_detection": True,
                "performance_optimized": True,
            },
        }

    def cleanup_cache(self, max_age_hours: int = 72):
        """Limpa cache antigo"""
        try:
            current_time = time.time()
            removed_count = 0

            for cache_file in self._cache_dir.glob("*.json"):
                try:
                    file_age = (current_time - cache_file.stat().st_mtime) / 3600
                    if file_age > max_age_hours:
                        cache_file.unlink()
                        removed_count += 1
                except Exception:
                    pass

            self.logger.info(
                f"Cache Excel cleanup: {removed_count} arquivos antigos removidos"
            )

        except Exception as e:
            self.logger.warning(f"Erro na limpeza do cache Excel: {e}")
